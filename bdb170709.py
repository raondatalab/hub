# -*- coding:utf-8 -*-
import time
# import sys
import traceback #에러스택확인
from selenium.webdriver.chrome.options import Options 
from selenium import webdriver   
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook
from time import localtime, strftime
from selenium.common.exceptions import TimeoutException

# VM 1~5
vm_no = str(4)

# reload(sys)
# sys.setdefaultencoding('utf-8')
def wait(sec):
    time.sleep(sec)

try:
	starttime_stamp = strftime('%Y-%m-%d %X\t', localtime())
	print (starttime_stamp)

	chromeOptions = webdriver.ChromeOptions()
	prefs = {"profile.managed_default_content_settings.images":2}
	chromeOptions.add_experimental_option("prefs",prefs)
	driver = webdriver.Chrome(chrome_options=chromeOptions)

	driver.get("http://raonserver1.cafe24.com")
	wait(3)
	elem = driver.find_element_by_id('login_id')
	elem.send_keys('admin')
	elem = driver.find_element_by_id('login_pw')
	elem.send_keys('qw12!@') 
	wait(1)
	elem.send_keys(Keys.RETURN)
	wait(3)
	driver.get("http://raonserver1.cafe24.com/bbs/board.php?bo_table=bdb&wr_id=" + vm_no)
	wait(1)
	status_elem = driver.find_element_by_class_name('wr_1')
	keyword_target_elem = driver.find_element_by_class_name('wr_2')
	location_elem = driver.find_element_by_class_name('wr_3')
	shop_elem = driver.find_element_by_class_name('wr_4')
	page_no_elem = driver.find_element_by_class_name('wr_5')
	# errpage_no_elem = driver.find_element_by_class_name('wr_6')
	status = status_elem.text
	keyword_target = keyword_target_elem.text
	location = location_elem.text
	shop = shop_elem.text
	page_no = page_no_elem.text
	# errpage_no = errpage_no_elem.text

	if (status == "서버대기"):
		print ("오더없음 서버대기")
		driver.quit()

	print ("실행시작")
	driver.get("http://raonserver1.cafe24.com/bbs/write.php?w=u&bo_table=bdb&wr_id=" + vm_no)
	wait(3)
	elem = driver.find_element_by_class_name('state')
	elem = elem.find_element_by_class_name('wr_1_3')
	elem.is_selected()
	elem.click()
	elem.send_keys(Keys.RETURN)

	area = "area" + vm_no
	print (area)
	if (area == "area1"):
		#41개
		area = ['서울특별시 종로구', '서울특별시 중구', '서울특별시 용산구', '서울특별시 성동구', '서울특별시 광진구', '서울특별시 동대문구', '서울특별시 중랑구', '서울특별시 성북구', '서울특별시 강북구', '서울특별시 도봉구', '서울특별시 노원구', '서울특별시 은평구', '서울특별시 서대문구', '서울특별시 마포구', '서울특별시 양천구', '서울특별시 강서구', '서울특별시 구로구', '서울특별시 금천구', '서울특별시 영등포구', '서울특별시 동작구', '서울특별시 관악구', '서울특별시 서초구', '서울특별시 강남구', '서울특별시 송파구', '서울특별시 강동구', '부산광역시 중구', '부산광역시 서구', '부산광역시 동구', '부산광역시 영도구', '부산광역시 부산진구', '부산광역시 동래구', '부산광역시 남구', '부산광역시 북구', '부산광역시 해운대구', '부산광역시 사하구', '부산광역시 금정구', '부산광역시 강서구', '부산광역시 연제구', '부산광역시 수영구', '부산광역시 사상구', '부산광역시 기장군']
	if (area == "area2"):
		#33개
		area = ['대전광역시 동구', '대전광역시 중구', '대전광역시 서구', '대전광역시 유성구', '대전광역시 대덕구', '대구광역시 중구', '대구광역시 동구', '대구광역시 서구', '대구광역시 남구', '대구광역시 북구', '대구광역시 수성구', '대구광역시 달서구', '대구광역시 달성군', '광주광역시 동구', '광주광역시 서구', '광주광역시 남구', '광주광역시 북구', '광주광역시 광산구', '울산광역시 중구', '울산광역시 남구', '울산광역시 동구', '울산광역시 북구', '울산광역시 울주군', '인천광역시 중구', '인천광역시 동구', '인천광역시 남구', '인천광역시 연수구', '인천광역시 남동구', '인천광역시 부평구', '인천광역시 계양구', '인천광역시 서구', '인천광역시 강화군', '인천광역시 옹진군']
	if (area == "area3"):
		#42개
		area = ['경기도 가평군', '경기도 고양시 덕양구', '경기도 고양시 일산동구', '경기도 고양시 일산서구', '경기도 과천시', '경기도 광명시', '경기도 광주시', '경기도 구리시', '경기도 군포시', '경기도 김포시', '경기도 남양주시', '경기도 동두천시', '경기도 부천시', '경기도 성남시 분당구', '경기도 성남시 수정구', '경기도 성남시 중원구', '경기도 수원시 권선구', '경기도 수원시 영통구', '경기도 수원시 장안구', '경기도 수원시 팔달구', '경기도 시흥시', '경기도 안산시 단원구', '경기도 안산시 상록구', '경기도 안성시', '경기도 안양시 동안구', '경기도 안양시 만안구', '경기도 양주시', '경기도 양평군', '경기도 여주시', '경기도 연천군', '경기도 오산시', '경기도 용인시 기흥구', '경기도 용인시 수지구', '경기도 용인시 처인구', '경기도 의왕시', '경기도 의정부시', '경기도 이천시', '경기도 파주시', '경기도 평택시', '경기도 포천시', '경기도 하남시', '경기도 화성시']
	if (area == "area4"):
		#50개
		area = ['강원도 강릉시', '강원도 고성군', '강원도 동해시', '강원도 삼척시', '강원도 속초시', '강원도 양구군', '강원도 양양군', '강원도 영월군', '강원도 원주시', '강원도 인제군', '강원도 정선군', '강원도 철원군', '강원도 춘천시', '강원도 태백시', '강원도 평창군', '강원도 홍천군', '강원도 화천군', '강원도 횡성군', '충청북도 괴산군', '충청북도 단양군', '충청북도 보은군', '충청북도 영동군', '충청북도 옥천군', '충청북도 음성군', '충청북도 제천시', '충청북도 증평군', '충청북도 진천군', '충청북도 청주시 상당구', '충청북도 청주시 서원구', '충청북도 청주시 청원구', '충청북도 청주시 흥덕구', '충청북도 충주시', '충청남도 계룡시', '충청남도 공주시', '충청남도 금산군', '충청남도 논산시', '충청남도 당진시', '충청남도 보령시', '충청남도 부여군', '충청남도 서산시', '충청남도 서천군', '충청남도 아산시', '충청남도 예산군', '충청남도 천안시 동남구', '충청남도 천안시 서북구', '충청남도 청양군', '충청남도 태안군', '충청남도 홍성군', '제주특별자치도 서귀포시', '제주특별자치도 제주시']
	if (area == "area5"):
		#83개
		area = ['경상북도 경산시', '경상북도 경주시', '경상북도 고령군', '경상북도 구미시', '경상북도 군위군', '경상북도 김천시', '경상북도 문경시', '경상북도 봉화군', '경상북도 상주시', '경상북도 성주군', '경상북도 안동시', '경상북도 영덕군', '경상북도 영양군', '경상북도 영주시', '경상북도 영천시', '경상북도 예천군', '경상북도 울릉군', '경상북도 울진군', '경상북도 의성군', '경상북도 청도군', '경상북도 청송군', '경상북도 칠곡군', '경상북도 포항시 남구', '경상북도 포항시 북구', '경상남도 거제시', '경상남도 거창군', '경상남도 고성군', '경상남도 김해시', '경상남도 남해군', '경상남도 밀양시', '경상남도 사천시', '경상남도 산청군', '경상남도 양산시', '경상남도 의령군', '경상남도 진주시', '경상남도 창녕군', '경상남도 창원시 마산합포구', '경상남도 창원시 마산회원구', '경상남도 창원시 성산구', '경상남도 창원시 의창구', '경상남도 창원시 진해구', '경상남도 통영시', '경상남도 하동군', '경상남도 함안군', '경상남도 함양군', '경상남도 합천군', '전라북도 고창군', '전라북도 군산시', '전라북도 김제시', '전라북도 남원시', '전라북도 무주군', '전라북도 부안군', '전라북도 순창군', '전라북도 완주군', '전라북도 익산시', '전라북도 임실군', '전라북도 장수군', '전라북도 전주시 덕진구', '전라북도 전주시 완산구', '전라북도 정읍시', '전라북도 진안군', '전라남도 강진군', '전라남도 고흥군', '전라남도 곡성군', '전라남도 광양시', '전라남도 구례군', '전라남도 나주시', '전라남도 담양군', '전라남도 목포시', '전라남도 무안군', '전라남도 보성군', '전라남도 순천시', '전라남도 신안군', '전라남도 여수시', '전라남도 영광군', '전라남도 영암군', '전라남도 완도군', '전라남도 장성군', '전라남도 장흥군', '전라남도 진도군', '전라남도 함평군', '전라남도 해남군', '전라남도 화순군']

	location_str = int(location)
	if (location_str != 0):
		area = area[location_str:]
	for area_i in area:
		new_wb = Workbook()
		new_ws = new_wb.active
		if (page_no != 0):
			page_no = page_no

		page_no = str(page_no)

		for get_list in range(500):
			page_no = int(page_no)
			page_no +=1
			page_no = str(page_no)
			try:
				driver.get("http://map.naver.com/?query='" + area_i + "' "  + keyword_target + "&page=" + page_no)
			except TimeoutException as ex:
   				print ("Timeout")
   				driver.refresh()
			wait(4)
			print (area_i + "page" + page_no)
			keyword_div = driver.find_element_by_class_name('search_result')
			get_list = keyword_div.find_elements_by_xpath('./ul/li')
			len_get_list = len(get_list)
			len_get_list = int(len_get_list)
			# print len_get_list
			# print type(len_get_list)
			wait(1)


			for get_list in get_list:
				# 타이틀
				title = get_list.find_element_by_xpath('./div[1]/dl/dt/a')
				buttons = keyword_div.find_elements_by_xpath('./ul/li')
				title.click()
				wait(3)
				# 새창정보
				keyword_sq = driver.find_element_by_class_name('default_dl')
				try:
					keyword_sq_err = driver.find_element_by_class_name('simplemodal-container')
					k_err = keyword_sq_err.text
					if (k_err[0] == '실'.decode('utf-8')):
						break
					if (k_err[0] == '서'.decode('utf-8')):
						wait(10)
						break
				except:
					pass
				address1 = keyword_sq.find_element_by_tag_name('dt')
				try:
					address2 = keyword_sq.find_element_by_class_name('info_road')
				except Exception as e:
					address2 = 0

				tel1 = keyword_sq.find_element_by_class_name('tel')
				cate = keyword_sq.find_element_by_class_name('cate')
				try:
					link1 = keyword_sq.find_element_by_xpath('./dd[3]/div[1]/a')
				except Exception as e:
					link1 = 0
				try:
					link2 = keyword_sq.find_element_by_xpath('./dd[3]/div[2]/a')
				except Exception as e:
					link2 = 0

				title = title.text
				address1 = address1.text
				if (address2 != 0):
					address2 = address2.text
				tel1 = tel1.text
				cate = cate.text
				if (link1 != 0):
					link1 = link1.text
				if (link2 != 0):
					link2 = link2.text

				max = new_ws.max_row
				new_ws.cell(row=1+max, column=1, value=title)
				new_ws.cell(row=1+max, column=2, value=address1)
				if (address2 != 0):
					new_ws.cell(row=1+max, column=3, value=address2[5:])
				new_ws.cell(row=1+max, column=4, value=tel1)
				new_ws.cell(row=1+max, column=5, value=cate)
				if (link1 != 0):
					new_ws.cell(row=1+max, column=6, value=link1)
				if (link2 != 0):
					new_ws.cell(row=1+max, column=7, value=link2)

				new_wb.save(area_i + " " + keyword_target + '.xlsx')
			if (len_get_list != 10):
				print (area_i + "END")
				break
		page_no = 0
		location_str = location_str + 1
	driver.quit()

except Exception as e:
	print (e)
	starttime_stamp = strftime('%Y-%m-%d %X\t', localtime())
	print ("err_time")
	print (starttime_stamp)
	print ("err_location")
	print (location_str)
	print ("err_page")
	print (page_no)
	# traceback.print_exc()  # 에러스택 정보를 stdout으로 print
	# errStr = traceback.format_exc() # 에러스택 정보를 String으로 반환
	driver.get("http://raonserver1.cafe24.com/bbs/write.php?w=u&bo_table=bdb&wr_id=" + vm_no)
	wait(3)
	elem = driver.find_element_by_class_name('wr_3')
	elem.clear()
	elem.send_keys(location_str)
	elem = driver.find_element_by_class_name('wr_5')
	elem.clear()
	elem.send_keys(page_no)
	elem.send_keys(Keys.RETURN)
	wait(1)
	driver.quit()